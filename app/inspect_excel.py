import pathlib
from openpyxl import load_workbook
import json

PATH = pathlib.Path(__file__).resolve().parent.parent / 'DostepnoscWTygodniach.xlsx'

if not PATH.exists():
    print('MISSING_FILE', PATH)
    raise SystemExit(1)

wb = load_workbook(PATH, data_only=True, read_only=True)
summary = {}
for sheet in wb.sheetnames:
    ws = wb[sheet]
    rows_iter = ws.iter_rows(min_row=1, max_row=10, values_only=True)
    headers = next(rows_iter, [])
    # znajdź pierwszą niepustą linię nagłówków
    while headers and all(h is None for h in headers):
        headers = next(rows_iter, [])
    sample = []
    for r in rows_iter:
        if any(v is not None for v in r):
            sample.append(r)
        if len(sample) >= 5:
            break
    summary[sheet] = {
        'headers': headers,
        'sample_rows': sample
    }

print(json.dumps(summary, ensure_ascii=False, indent=2))
